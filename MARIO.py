import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# CONSULTA EN SQL
server = 'MP-VW-DB-017'
database = 'Producto2'
conn = pyodbc.connect(f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database}')
consulta_sql = """
 
    SELECT 
    DESCRIPCION,
    MONTO,
    CORTE,
    COUNT(MONTO) AS CantidadMontosIguales,
    SUM(CAST(MONTO AS DECIMAL(10, 2))) AS SumaMontosIguales,
    Mes_Cargo AS MesCargo,
    DATEFROMPARTS(YEAR(GETDATE()), Mes_Cargo, CORTE) AS MesConsulta
FROM NB_Cargo_Anualidad_lagm
--WHERE Mes_Cargo = MONTH(GETDATE())
GROUP BY DESCRIPCION, MONTO, CORTE, Mes_Cargo
ORDER BY CORTE, DESCRIPCION, MONTO;

"""

# Execute SQL query and create a DataFrame
cursor = conn.cursor()
cursor.execute(consulta_sql)
rows = cursor.fetchall()
column_names = [column[0] for column in cursor.description]
consulta = pd.DataFrame.from_records(rows, columns=column_names)

# Define the Excel file path
excel_file_path = r'\\celerrabr\Sesiones\D7627S066\Comum_S066\CRM y MKT\CRM\Mario Chavez\Anualidades prueba\Bitacora Anualidades PRUEBA.xlsx'
nombre_pagina = 'Consulta'

# Load the existing workbook and sheet
workbook = load_workbook(excel_file_path)
nueva_hoja = workbook[nombre_pagina]

# Clear the existing data in the sheet
nueva_hoja.delete_rows(2, nueva_hoja.max_row)

# Write the new DataFrame to the Excel sheet
for r_idx, row in enumerate(dataframe_to_rows(consulta, index=False, header=True), start=2):
    for c_idx, value in enumerate(row, start=1):
        nueva_hoja.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook to the specified path
workbook.save(excel_file_path)
print('Excel file updated successfully at:', excel_file_path)



